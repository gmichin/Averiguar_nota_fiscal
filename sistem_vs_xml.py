import os
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
import pandas as pd
from pathlib import Path
import chardet
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def detectar_encoding(arquivo):
    """Detectar a codificação do arquivo"""
    with open(arquivo, 'rb') as f:
        resultado = chardet.detect(f.read())
    return resultado['encoding']

def converter_para_float(valor):
    """Converter valor para float, tratando vírgulas como separador decimal"""
    if pd.isna(valor) or valor == '':
        return 0.0
    try:
        if isinstance(valor, str):
            valor = valor.strip().replace('.', '').replace(',', '.')
        return float(valor)
    except (ValueError, TypeError):
        return 0.0

def converter_para_int(valor):
    """Converter valor para inteiro, tratando casos especiais"""
    if pd.isna(valor) or valor == '':
        return 0
    try:
        if isinstance(valor, str):
            valor = valor.strip()
            valor = ''.join(filter(lambda x: x.isdigit() or x in '.,', valor))
            if '.' in valor or ',' in valor:
                valor_float = converter_para_float(valor)
                return int(valor_float)
        return int(float(valor))
    except (ValueError, TypeError):
        return 0

def formatar_data(data_xml):
    """Converte a data do formato XML para formato legível"""
    try:
        data_str = data_xml.split('-03:00')[0] if '-03:00' in data_xml else data_xml.split('-04:00')[0] if '-04:00' in data_xml else data_xml
        
        if 'T' in data_str:
            data_dt = datetime.strptime(data_str, '%Y-%m-%dT%H:%M:%S')
        else:
            data_dt = datetime.strptime(data_str, '%Y-%m-%d %H:%M:%S')
        
        return data_dt.strftime('%d/%m/%Y %H:%M')
    except Exception:
        return data_xml

def extrair_numero_nota_do_nome(nome_arquivo):
    """Extrai o número da nota fiscal do nome do arquivo"""
    try:
        # Remove extensão .xml
        nome_sem_ext = nome_arquivo.replace('.xml', '').replace('.XML', '')
        # Extrai apenas números
        numeros = ''.join(filter(str.isdigit, nome_sem_ext))
        if numeros:
            return int(numeros)
    except:
        pass
    return None

def verificar_cancelamento_intempestivo(caminhos_recusado, nfe_str):
    """Verifica se há arquivo na pasta recusado e se contém a mensagem de cancelamento intempestivo"""
    padrao_arquivo = f"*{nfe_str}*.txt"
    
    for caminho_recusado in caminhos_recusado:
        if not os.path.exists(caminho_recusado):
            continue
            
        try:
            for arquivo in Path(caminho_recusado).glob(padrao_arquivo):
                try:
                    encoding = detectar_encoding(arquivo)
                    with open(arquivo, 'r', encoding=encoding) as f:
                        conteudo = f.read()
                    
                    # Verificar se contém a mensagem específica
                    if any(msg in conteudo for msg in [
                        "501 : Rejeição: Pedido de Cancelamento intempestivo",
                        "493 : Rejeição: Evento não atende o Schema XML específico", 
                        "221 : Rejeição: Confirmado o recebimento da NF-e pelo destinatário",
                        "241 : Rejeição: Um número da faixa já foi utilizado"
                    ]):
                        return True
                except Exception:
                    continue
        except Exception:
            continue
    
    return False

def carregar_arquivos_can_rapido(caminhos_eventos):
    """Carrega lista de arquivos .can de forma rápida"""
    arquivos_can = set()
    for caminho_evento in caminhos_eventos:
        if os.path.exists(caminho_evento):
            try:
                # Listar arquivos .can uma única vez
                for arquivo in os.listdir(caminho_evento):
                    if arquivo.lower().endswith('.can'):
                        arquivos_can.add(arquivo.lower())
            except Exception as e:
                print(f"⚠️ Erro ao acessar eventos {caminho_evento}: {e}")
    return arquivos_can

def verificar_inutilizacao_nota_nao_autorizada(caminhos_eventos, nfe_num):
    """Verifica se há arquivo .inu com NOTA NAO AUTORIZADA para a nota fiscal"""
    nfe_str = str(nfe_num).zfill(8)
    padrao_arquivo = f"*{nfe_str}*.inu"
    
    for caminho_evento in caminhos_eventos:
        if not os.path.exists(caminho_evento):
            continue
            
        try:
            for arquivo in Path(caminho_evento).glob(padrao_arquivo):
                try:
                    encoding = detectar_encoding(arquivo)
                    with open(arquivo, 'r', encoding=encoding) as f:
                        conteudo = f.read()
                    
                    # Verificar se contém a mensagem específica
                    if any(msg in conteudo for msg in [
                        '<xJust>NOTA NAO AUTORIZADA</xJust>',
                        '<xJust>NOTA NAO APARECE NO SEFAZ</xJust>',
                        '<xServ>INUTILIZAR</xServ>',
                        '<xJust>ERRO NO SEFAZ................</xJust>',
                        '<xJust>MERCADO NAO QUIS RECEBER</xJust>',
                        '<xJust>MERCADORIA FOI DUAS VEZES NO DIA</xJust>',
                        '<xJust>ERRO NA PESAGEM</xJust>',
                        '<xJust>NAO APARECE NO SEFAZ....</xJust>',
                        '<xJust>IMPOSTO ERRADO......</xJust>',
                        '<xJust>FORA DE HORARIO....</xJust>',
                        '<xJust>CARRO QUEBROU.........</xJust>',
                        '<xJust>VENDEDORA DIGITOU QNT ERRADA...</xJust>',
                        '<xJust>CLIENTE EM INVENTARIO</xJust>'
                    ]):
                        return True
                except Exception:
                    continue
        except Exception:
            continue
    
    return False

def extrair_dados_basicos_xml(caminho_completo):
    """Extrai apenas dados essenciais do XML (muito mais rápido)"""
    try:
        encoding = detectar_encoding(caminho_completo)
        
        # Ler apenas os primeiros 10KB do arquivo (onde geralmente estão os dados principais)
        with open(caminho_completo, 'r', encoding=encoding) as file:
            conteudo = file.read(10000)  # Lê apenas os primeiros 10KB
        
        # Verificar se é nota de venda rapidamente
        if '<natOp>VENDA</natOp>' not in conteudo:
            return None
        
        # Extrair dados usando busca em texto (mais rápido que XML parser)
        def extrair_tag(tag, texto):
            tag_abertura = f'<{tag}>'
            tag_fechamento = f'</{tag}>'
            inicio = texto.find(tag_abertura)
            if inicio != -1:
                inicio += len(tag_abertura)
                fim = texto.find(tag_fechamento, inicio)
                if fim != -1:
                    return texto[inicio:fim]
            return None
        
        cnf = extrair_tag('cNF', conteudo)
        nnf = extrair_tag('nNF', conteudo)
        vnf = extrair_tag('vNF', conteudo)
        dh_emi = extrair_tag('dhEmi', conteudo)
        
        if cnf and nnf and vnf and dh_emi:
            return {
                'Romaneio': int(cnf),
                'NF-E': int(nnf),
                'Valor XML': float(vnf),
                'DATA': dh_emi
            }
        
        return None
        
    except Exception:
        return None

def processar_xml_completo(caminho_completo, dados_basicos, arquivos_can, caminhos_recusado, caminhos_eventos):
    """Processa um arquivo XML completo e retorna os dados"""
    try:
        nfe_num = dados_basicos['NF-E']
        nfe_str = str(nfe_num).zfill(8)
        nome_can = f"{nfe_str}.can"
        
        # PRIMEIRO: Verificar se a nota foi inutilizada com "NOTA NAO AUTORIZADA"
        if verificar_inutilizacao_nota_nao_autorizada(caminhos_eventos, nfe_num):
            print(f"⚠️ Nota {nfe_num} inutilizada (NÃO AUTORIZADA) - removendo da lista")
            return None
        
        # SEGUNDO: Verificar se existe arquivo .can
        if nome_can.lower() in arquivos_can:
            # Verificar se há cancelamento intempestivo
            if verificar_cancelamento_intempestivo(caminhos_recusado, nfe_str):
                return {
                    'CF': 'VENDA',
                    'Romaneio': dados_basicos['Romaneio'],
                    'NF-E': nfe_num,
                    'Valor XML': dados_basicos['Valor XML'],
                    'DATA': formatar_data(dados_basicos['DATA']),
                    'OBS': 'Cancelamento Intempestivo'
                }
            else:
                return None  # Nota cancelada normalmente
        else:
            # Nota não cancelada
            return {
                'CF': 'VENDA',
                'Romaneio': dados_basicos['Romaneio'],
                'NF-E': nfe_num,
                'Valor XML': dados_basicos['Valor XML'],
                'DATA': formatar_data(dados_basicos['DATA'])
            }
    
    except Exception as e:
        print(f"⚠️ Erro ao processar {os.path.basename(caminho_completo)}: {e}")
    
    return None

def buscar_xml_por_data():
    """Processa XMLs de notas fiscais por período - VERSÃO ULTRARRÁPIDA"""
    print("=== PROCESSADOR DE NOTAS FISCAIS (VERSÃO RÁPIDA) ===")
    data_inicial_str = input("Digite a data inicial (DD/MM/AAAA): ")
    data_final_str = input("Digite a data final (DD/MM/AAAA): ")
    
    try:
        data_inicial = datetime.strptime(data_inicial_str, "%d/%m/%Y").date()
        data_final = datetime.strptime(data_final_str, "%d/%m/%Y").date()
        
        # INCLUIR TODOS OS DIAS ENTRE AS DATAS
        dias_periodo = (data_final - data_inicial).days + 1
        print(f"📅 Período: {data_inicial_str} a {data_final_str} ({dias_periodo} dias)")
        
    except ValueError:
        print("❌ Formato de data inválido!")
        return None
    
    # Lista de caminhos - INCLUINDO PASTAS ENVIADO
    caminhos_xml = [
        r"S:\hor\nfe",
        r"S:\hor\nfe2",
        r"S:\hor\nfe\enviado",
        r"S:\hor\nfe2\enviado"
    ]
    
    caminhos_eventos = [
        r"S:\hor\nfe\eventos",
        r"S:\hor\nfe2\eventos"
    ]
    
    caminhos_recusado = [
        r"S:\hor\nfe\recusado",
        r"S:\hor\nfe2\recusado"
    ]
    
    # Verificar diretórios
    diretorios_existentes = [c for c in caminhos_xml if os.path.exists(c)]
    if not diretorios_existentes:
        print("❌ Nenhum diretório encontrado!")
        return None
    
    print("⏳ Carregando arquivos .can...")
    arquivos_can = carregar_arquivos_can_rapido(caminhos_eventos)
    print(f"📄 {len(arquivos_can)} arquivos .can carregados")
    
    print("⏳ Buscando arquivos XML no período...")
    
    dados_nfe = []
    total_arquivos = 0
    arquivos_no_periodo = 0
    notas_processadas = 0
    
    # PRIMEIRA ETAPA: Filtrar arquivos pela data de modificação (MUITO RÁPIDO)
    arquivos_para_processar = []
    arquivos_unicos = {}
    
    for caminho_xml in diretorios_existentes:
        print(f"🔍 Escaneando {caminho_xml}...")
        
        try:
            # Listar arquivos uma única vez
            with os.scandir(caminho_xml) as entries:
                arquivos_lista = [entry for entry in entries if entry.is_file() and entry.name.lower().endswith('.xml')]
            
            total_arquivos += len(arquivos_lista)
            
            # Filtrar pela data de modificação do arquivo
            for entry in arquivos_lista:
                # Verificar data de modificação do arquivo
                data_modificacao = datetime.fromtimestamp(entry.stat().st_mtime).date()
                
                # Se a data de modificação estiver dentro do período
                if data_inicial <= data_modificacao <= data_final:
                    # Evitar duplicatas pelo nome do arquivo
                    if entry.name not in arquivos_unicos:
                        arquivos_unicos[entry.name] = {
                            'caminho': entry.path,
                            'data_mod': data_modificacao
                        }
                        arquivos_no_periodo += 1
                        
        except Exception as e:
            print(f"⚠️ Erro em {caminho_xml}: {e}")
    
    print(f"📊 Total de arquivos XML encontrados: {total_arquivos}")
    print(f"📅 Arquivos únicos no período (pela data modificação): {arquivos_no_periodo}")
    
    if arquivos_no_periodo == 0:
        print("❌ Nenhum arquivo no período especificado.")
        return None
    
    # SEGUNDA ETAPA: Processar apenas os arquivos do período (leitura rápida)
    print("⏳ Extraindo dados dos XMLs...")
    
    for i, (nome_arquivo, info) in enumerate(arquivos_unicos.items(), 1):
        if i % 50 == 0:
            print(f"📦 Processados {i}/{arquivos_no_periodo} arquivos...")
        
        # Extrair dados básicos rapidamente
        dados_basicos = extrair_dados_basicos_xml(info['caminho'])
        
        if dados_basicos:
            # Verificar se a data do XML está dentro do período
            try:
                data_xml_str = dados_basicos['DATA']
                if 'T' in data_xml_str:
                    data_xml = datetime.strptime(data_xml_str.split('T')[0], '%Y-%m-%d').date()
                else:
                    data_xml = datetime.strptime(data_xml_str.split()[0], '%Y-%m-%d').date()
                
                if data_inicial <= data_xml <= data_final:
                    # Processar dados completos
                    dados = processar_xml_completo(info['caminho'], dados_basicos, arquivos_can, caminhos_recusado, caminhos_eventos)
                    if dados:
                        dados_nfe.append(dados)
                        notas_processadas += 1
            except:
                # Se não conseguir extrair data, usa o arquivo mesmo assim
                dados = processar_xml_completo(info['caminho'], dados_basicos, arquivos_can, caminhos_recusado, caminhos_eventos)
                if dados:
                    dados_nfe.append(dados)
                    notas_processadas += 1
    
    print(f"\n📊 RESUMO FINAL:")
    print(f"📄 Arquivos no período: {arquivos_no_periodo}")
    print(f"✅ Notas processadas: {notas_processadas}")
    if dados_nfe:
        print(f"💰 Valor total: R$ {sum(d['Valor XML'] for d in dados_nfe):,.2f}")
    
    if dados_nfe:
        df_resultado = pd.DataFrame(dados_nfe)
        df_resultado = df_resultado.sort_values('NF-E')
        return df_resultado
    else:
        print("⚠️ Nenhuma nota fiscal de VENDA encontrada no período.")
        return None
    
def processar_faturamento_bruto():
    """Processa arquivos CSV para faturamento bruto"""
    caminho_fechamento = r"S:\hor\excel\fechamento-20260401-20260417.csv"
    caminho_cancelados = r"S:\hor\arquivos\gustavo\can.csv"
    caminho_historico = r"S:\hor\excel\20260401.csv"
    
    try:
        encoding_fechamento = detectar_encoding(caminho_fechamento)
        df_principal = pd.read_csv(caminho_fechamento, encoding=encoding_fechamento, sep=';', decimal=',')
        
        if df_principal.empty:
            return None
        
        df_principal.columns = df_principal.columns.str.strip().str.upper()
        
        colunas_necessarias = ['LOJA', 'RAZAO', 'GRUPO', 'ROMANEIO', 'NF-E', 'DATA', 
                              'VENDEDOR', 'CODPRODUTO', 'GRUPO PRODUTO', 'DESCRICAO', 'PRECO VENDA']
        colunas_existentes = [col for col in colunas_necessarias if col in df_principal.columns]
        
        if not colunas_existentes:
            return None
        
        df_principal = df_principal[colunas_existentes]
        
        print("⏳ Convertendo colunas para numérico...")
        
        if 'ROMANEIO' in df_principal.columns:
            df_principal['ROMANEIO'] = df_principal['ROMANEIO'].apply(converter_para_int)
        
        if 'NF-E' in df_principal.columns:
            df_principal['NF-E'] = df_principal['NF-E'].apply(converter_para_int)
        
        if 'CODPRODUTO' in df_principal.columns:
            df_principal['CODPRODUTO'] = df_principal['CODPRODUTO'].apply(converter_para_int)
        
        df_principal['PRECO VENDA'] = df_principal['PRECO VENDA'].apply(converter_para_float)
        df_principal = df_principal[df_principal['PRECO VENDA'] >= 0]
        
        try:
            encoding_cancelados = detectar_encoding(caminho_cancelados)
            df_cancelados = pd.read_csv(caminho_cancelados, skiprows=2, encoding=encoding_cancelados, sep=';')
            
            if len(df_cancelados.columns) > 0:
                nfes_cancelados = df_cancelados.iloc[:, 0].dropna().apply(converter_para_int).unique()
                df_principal = df_principal[~df_principal['NF-E'].isin(nfes_cancelados)]
        except Exception:
            pass
        
        try:
            encoding_historico = detectar_encoding(caminho_historico)
            df_historico = pd.read_csv(caminho_historico, encoding=encoding_historico, sep=';')
            df_historico.columns = df_historico.columns.str.strip().str.upper()
            
            colunas_historico = ['ROMANEIO', 'NOTA FISCAL', 'PRODUTO', 'HISTORICO', 'PESO']
            colunas_existentes_hist = [col for col in colunas_historico if col in df_historico.columns]
            
            if colunas_existentes_hist:
                df_historico = df_historico[colunas_existentes_hist]
                
                if 'ROMANEIO' in df_historico.columns:
                    df_historico['ROMANEIO'] = df_historico['ROMANEIO'].apply(converter_para_int)
                if 'NOTA FISCAL' in df_historico.columns:
                    df_historico['NOTA FISCAL'] = df_historico['NOTA FISCAL'].apply(converter_para_int)
                if 'PRODUTO' in df_historico.columns:
                    df_historico['PRODUTO'] = df_historico['PRODUTO'].apply(converter_para_int)
                
                df_principal['PESO'] = 0.0
                
                linhas_para_remover = []
                indices_com_peso = []
                
                for idx, row_principal in df_principal.iterrows():
                    mask = (
                        (df_historico['ROMANEIO'] == row_principal['ROMANEIO']) &
                        (df_historico['NOTA FISCAL'] == row_principal['NF-E']) &
                        (df_historico['PRODUTO'] == row_principal['CODPRODUTO'])
                    )
                    
                    correspondencias = df_historico[mask]
                    
                    if not correspondencias.empty:
                        historico_valor = pd.to_numeric(correspondencias['HISTORICO'].iloc[0], errors='coerce')
                        
                        if historico_valor == 68:
                            linhas_para_remover.append(idx)
                        elif historico_valor == 51 and 'PESO' in correspondencias.columns:
                            peso_valor = converter_para_float(correspondencias['PESO'].iloc[0])
                            indices_com_peso.append((idx, peso_valor))
                
                if linhas_para_remover:
                    df_principal = df_principal.drop(linhas_para_remover)
                
                for idx, peso in indices_com_peso:
                    if idx in df_principal.index:
                        df_principal.at[idx, 'PESO'] = peso
                        
        except Exception:
            pass
        
        df_principal['PESO'] = df_principal['PESO'].apply(converter_para_float)
        df_principal['FAT BRUTO'] = df_principal['PRECO VENDA'] * df_principal['PESO']
        
        print(f"✅ {len(df_principal)} linhas processadas")
        return df_principal
        
    except Exception as e:
        print(f"❌ Erro no processamento: {e}")
        return None

def criar_tabela_excel_com_formatacao(df_xml, df_faturamento):
    """Cria arquivo Excel com tabelas reais inseridas e linhas de totais"""
    downloads_path = str(Path.home() / "Downloads")
    caminho_excel = os.path.join(downloads_path, "SISTEMA_X_XML.xlsx")
    
    # Criar workbook
    wb = Workbook()
    
    # Remover sheet padrão vazio
    wb.remove(wb.active)
    
    try:
        # ABA 1: NOTAS FISCAIS
        if df_xml is not None and not df_xml.empty:
            ws_nf = wb.create_sheet("Notas Fiscais")
            
            # Adicionar cabeçalhos
            cabecalhos = list(df_xml.columns)
            for col_num, cabecalho in enumerate(cabecalhos, 1):
                col_letra = get_column_letter(col_num)
                ws_nf[f'{col_letra}1'] = cabecalho
            
            # Adicionar dados
            for row_num, row_data in enumerate(df_xml.values, 2):
                for col_num, value in enumerate(row_data, 1):
                    col_letra = get_column_letter(col_num)
                    ws_nf[f'{col_letra}{row_num}'] = value
            
            # Adicionar linha de totais
            total_row = len(df_xml) + 2
            ws_nf[f'A{total_row}'] = 'TOTAL'
            
            # Encontrar coluna do Valor XML
            valor_xml_col = None
            for idx, col_name in enumerate(df_xml.columns, 1):
                if 'VALOR XML' in col_name.upper():
                    valor_xml_col = get_column_letter(idx)
                    break
            
            if valor_xml_col:
                # Calcular total
                total_valor_xml = df_xml['Valor XML'].sum()
                ws_nf[f'{valor_xml_col}{total_row}'] = total_valor_xml
                
                # Formatar a célula de total
                ws_nf[f'{valor_xml_col}{total_row}'].font = Font(bold=True)
                ws_nf[f'{valor_xml_col}{total_row}'].alignment = Alignment(horizontal='right')
                ws_nf[f'A{total_row}'].font = Font(bold=True)
            
            # Criar tabela (sem incluir a linha de totais)
            max_row = len(df_xml) + 1
            max_col = len(df_xml.columns)
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            
            tab_nf = Table(displayName="TabelaNotasFiscais", ref=ref)
            tab_nf.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws_nf.add_table(tab_nf)
            
            # Ajustar largura das colunas
            for col in ws_nf.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_nf.column_dimensions[col_letter].width = min(adjusted_width, 50)
            
            print(f"✅ Tabela 'Notas Fiscais' criada com {len(df_xml)} registros")
        
        # ABA 2: FATURAMENTO BRUTO
        if df_faturamento is not None and not df_faturamento.empty:
            ws_fat = wb.create_sheet("Faturamento Bruto")
            
            # Adicionar cabeçalhos
            cabecalhos = list(df_faturamento.columns)
            for col_num, cabecalho in enumerate(cabecalhos, 1):
                col_letra = get_column_letter(col_num)
                ws_fat[f'{col_letra}1'] = cabecalho
            
            # Adicionar dados
            for row_num, row_data in enumerate(df_faturamento.values, 2):
                for col_num, value in enumerate(row_data, 1):
                    col_letra = get_column_letter(col_num)
                    ws_fat[f'{col_letra}{row_num}'] = value
            
            # Adicionar linha de totais
            total_row = len(df_faturamento) + 2
            ws_fat[f'A{total_row}'] = 'TOTAL'
            
            # Encontrar coluna do FAT BRUTO
            fat_bruto_col = None
            for idx, col_name in enumerate(df_faturamento.columns, 1):
                if 'FAT BRUTO' in col_name.upper():
                    fat_bruto_col = get_column_letter(idx)
                    break
            
            if fat_bruto_col:
                # Calcular total
                total_fat_bruto = df_faturamento['FAT BRUTO'].sum()
                ws_fat[f'{fat_bruto_col}{total_row}'] = total_fat_bruto
                
                # Formatar a célula de total
                ws_fat[f'{fat_bruto_col}{total_row}'].font = Font(bold=True)
                ws_fat[f'{fat_bruto_col}{total_row}'].alignment = Alignment(horizontal='right')
                ws_fat[f'A{total_row}'].font = Font(bold=True)
            
            # Criar tabela (sem incluir a linha de totais)
            max_row = len(df_faturamento) + 1
            max_col = len(df_faturamento.columns)
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            
            tab_fat = Table(displayName="TabelaFaturamento", ref=ref)
            tab_fat.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws_fat.add_table(tab_fat)
            
            # Ajustar largura das colunas
            for col in ws_fat.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_fat.column_dimensions[col_letter].width = min(adjusted_width, 50)
            
            print(f"✅ Tabela 'Faturamento Bruto' criada com {len(df_faturamento)} registros")
        
        # Salvar arquivo
        wb.save(caminho_excel)
        print(f"✅ Arquivo salvo com tabelas e totais inseridos: {caminho_excel}")
        return True
        
    except Exception as e:
        print(f"❌ Erro ao criar tabelas: {e}")
        return False

def main():
    """Função principal"""
    print("=== SISTEMA X XML COM TABELAS E TOTAIS (VERSÃO OTIMIZADA) ===")
    print("1. Processar XMLs de Notas Fiscais")
    print("2. Processar Faturamento Bruto")
    print("3. Processar Ambos")
    
    opcao = input("Escolha uma opção (1/2/3): ").strip()
    
    df_xml = None
    df_faturamento = None
    
    if opcao in ['1', '3']:
        print("\n📁 Processando XMLs...")
        df_xml = buscar_xml_por_data()
    
    if opcao in ['2', '3']:
        print("\n📊 Processando Faturamento...")
        df_faturamento = processar_faturamento_bruto()
    
    if df_xml is not None or df_faturamento is not None:
        sucesso = criar_tabela_excel_com_formatacao(df_xml, df_faturamento)
        
        if sucesso:
            # Estatísticas
            if df_xml is not None and not df_xml.empty:
                total_valor_xml = df_xml['Valor XML'].sum()
                print(f"\n📊 Notas Fiscais: {len(df_xml)} registros | Total: R$ {total_valor_xml:,.2f}")
            
            if df_faturamento is not None and not df_faturamento.empty:
                total_fat_bruto = df_faturamento['FAT BRUTO'].sum() if 'FAT BRUTO' in df_faturamento.columns else 0
                print(f"📊 Faturamento Bruto: {len(df_faturamento)} registros | Total: R$ {total_fat_bruto:,.2f}")
            
            print("\n💡 DICA: Ao abrir o Excel, você verá:")
            print("   • Tabelas formatadas com filtros automáticos")
            print("   • Linha de totais abaixo de cada tabela")
            print("   • Formatação em negrito para os totais")
        else:
            print("❌ Erro ao criar arquivo com tabelas.")
    else:
        print("❌ Nenhum dado foi processado.")

if __name__ == "__main__":
    # Verificar dependências
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("⏳ Instalando openpyxl...")
        import subprocess
        subprocess.check_call(["pip", "install", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment
    
    try:
        import chardet
    except ImportError:
        print("⏳ Instalando chardet...")
        import subprocess
        subprocess.check_call(["pip", "install", "chardet"])
        import chardet
    
    main()
    input("\nPressione Enter para sair...")